"""Unit tests for `src/modules/reports/service.py`.

Covers:
  - `ReportService.generate` persists narrative + token counts on completion
  - `ReportService.generate` marks the run as failed when Claude errors
  - `ReportService.generate` streams deltas in order
  - `get_artifact(csv)` returns a valid CSV byte string
  - `get_artifact(xlsx)` returns a valid XLSX byte string
  - `get_artifact(pdf)` returns a PDF byte string (starts with %PDF-)
  - `_compute_cost_usd` returns the right dollar amount for Sonnet 4.5

The Anthropic SDK is mocked entirely — no real network calls. The
session is mocked at the level of `execute()` and `flush()`/`commit()`
so the persistence half of `generate` can be exercised without a DB.
"""

from __future__ import annotations

import csv
import io
import uuid
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import pytest

from src.modules.reports.schemas import ReportType
from src.modules.reports.service import (
    ReportService,
    _compute_cost_usd,
    _flatten_aggregate_for_table,
    get_artifact,
    render_csv,
    render_pdf,
    render_xlsx,
)

# Importing `src.main` here (instead of `from src.modules.reports.service`
# alone) makes sure every ORM model is registered with the declarative
# `Base.metadata` before any query runs — otherwise SQLAlchemy raises
# "Mapper[...] failed to locate a name ('Agency')" because `relationship()`
# strings resolve lazily.
_ = __import__("src.main", fromlist=["app"])


# --------------------------------------------------------------------------
# Test fixtures
# --------------------------------------------------------------------------
class _FakeStreamCM:
    """Async context manager that yields a fake SDK stream."""

    def __init__(self, events: list[Any]) -> None:
        self._events = events

    async def __aenter__(self) -> _FakeStream:
        return _FakeStream(self._events)

    async def __aexit__(self, exc_type, exc, tb) -> None:
        return None


class _FakeStream:
    """Mimics the AsyncMessageStream async-iterable protocol."""

    def __init__(self, events: list[Any]) -> None:
        self._events = list(events)
        self._iter = iter(self._events)

    def __aiter__(self) -> _FakeStream:
        return self

    async def __anext__(self) -> Any:
        try:
            return next(self._iter)
        except StopIteration as exc:
            raise StopAsyncIteration from exc


def _msg_start(input_tokens: int = 100) -> Any:
    """Build a fake MessageStartEvent with the given input_tokens.

    Uses `SimpleNamespace` (not `MagicMock`) so `hasattr(...)` returns
    `False` for unrelated attributes — the service's duck-typed event
    dispatch relies on that to take the right branch.
    """
    from types import SimpleNamespace

    return SimpleNamespace(
        message=SimpleNamespace(usage=SimpleNamespace(input_tokens=input_tokens))
    )


def _text_delta(text: str) -> Any:
    """Build a fake ContentBlockDeltaEvent with a text delta."""
    from types import SimpleNamespace

    return SimpleNamespace(
        delta=SimpleNamespace(type="text_delta", text=text),
    )


def _msg_delta(output_tokens: int = 200) -> Any:
    """Build a fake MessageDeltaEvent with the given output_tokens."""
    from types import SimpleNamespace

    return SimpleNamespace(usage=SimpleNamespace(output_tokens=output_tokens))


def _msg_stop() -> Any:
    """Build a fake MessageStopEvent — no fields read by our code."""
    from types import SimpleNamespace

    return SimpleNamespace()


def _fake_session_with_run(run_id: uuid.UUID | None = None) -> AsyncMock:
    """Build a fake AsyncSession that captures `add()` + `flush()`.

    The session's `commit()` is also an AsyncMock so `generate()`
    can call it without raising. We don't actually round-trip a row —
    we only assert on the side-effects.

    `execute()` is wired to return empty result objects so the aggregator
    queries (`select(...)`, `func.count(...)`) don't crash. Tests that
    need a non-empty aggregate should subclass and override.

    Note: `session.add` is a SYNC `MagicMock` (not AsyncMock) because
    SQLAlchemy's `Session.add()` is synchronous. The async-mock it
    inherits from by default would swallow the call without capturing
    the row, and our side-effect lambda wouldn't fire.
    """
    from unittest.mock import MagicMock as _MM  # noqa: N814

    session = AsyncMock()
    if run_id is None:
        run_id = uuid.uuid4()

    captured: dict[str, Any] = {}

    def _capture(obj: Any) -> None:
        captured["run"] = obj

    # Use a sync MagicMock for `add` — `add()` is synchronous.
    session.add = _MM(side_effect=_capture)
    session.flush = AsyncMock(
        side_effect=lambda: setattr(captured["run"], "id", run_id)
    )
    session.commit = AsyncMock()

    # `session.execute(stmt)` returns a Result-like object. The aggregator
    # either calls `.one()` (visit_summary) or `.scalars()` (lists) or
    # `.scalar()` (counts). Empty defaults keep it crash-free.
    empty_result = _MM()
    empty_result.one.return_value = _MM(
        total=0,
        completed=0,
        total_seconds=0,
        with_gps=0,
    )
    empty_scalars = _MM()
    empty_scalars.all.return_value = []
    empty_result.scalars.return_value = empty_scalars
    empty_result.scalar.return_value = 0

    session.execute = AsyncMock(return_value=empty_result)
    # Expose `captured` for tests that want to read the row directly.
    session._captured = captured  # type: ignore[attr-defined]
    return session


def _patch_anthropic_client(monkeypatch, events: list[Any]) -> None:
    """Patch the Anthropic client's `messages.stream` to emit `events`."""
    fake_client = MagicMock()
    fake_client.messages.stream = MagicMock(return_value=_FakeStreamCM(events))
    monkeypatch.setattr(
        "anthropic.AsyncAnthropic",
        lambda **kwargs: fake_client,
    )


# --------------------------------------------------------------------------
# ReportService.generate
# --------------------------------------------------------------------------
class TestReportServiceGenerate:
    @pytest.mark.asyncio
    async def test_persists_narrative_on_completion(self, monkeypatch) -> None:
        """Streaming succeeds → row.status='completed', narrative set, tokens counted."""
        run_id = uuid.uuid4()
        session = _fake_session_with_run(run_id)
        events = [
            _msg_start(input_tokens=42),
            _text_delta("Hello "),
            _text_delta("world"),
            _msg_delta(output_tokens=10),
            _msg_stop(),
        ]
        _patch_anthropic_client(monkeypatch, events)

        service = ReportService()
        events_out = []
        async for ev in service.generate(
            session,
            agency_id=uuid.uuid4(),
            user_id=uuid.uuid4(),
            report_type=ReportType.VISIT_SUMMARY,
            params={"date_from": "2026-07-01"},
        ):
            events_out.append(ev)

        # First event is the run_meta; last is the final.
        assert events_out[0].kind == "run_meta"
        assert events_out[-1].kind == "final"
        assert events_out[-1].total_tokens == 52  # 42 + 10

        run = session._captured["run"]
        assert run.status == "completed"
        assert run.narrative == "Hello world"
        assert run.input_tokens == 42
        assert run.output_tokens == 10
        assert run.error is None
        assert run.completed_at is not None
        session.commit.assert_awaited()

    @pytest.mark.asyncio
    async def test_marks_failed_on_claude_error(self, monkeypatch) -> None:
        """SDK raises → row.status='failed', error captured, error frame yielded."""
        run_id = uuid.uuid4()
        session = _fake_session_with_run(run_id)

        # Stream raises on enter.
        class _BoomCM(_FakeStreamCM):
            async def __aenter__(self):  # type: ignore[override]
                raise RuntimeError("boom")

        fake_client = MagicMock()
        fake_client.messages.stream = MagicMock(return_value=_BoomCM([]))
        monkeypatch.setattr(
            "anthropic.AsyncAnthropic", lambda **kwargs: fake_client
        )

        service = ReportService()
        events_out = []
        async for ev in service.generate(
            session,
            agency_id=uuid.uuid4(),
            user_id=uuid.uuid4(),
            report_type=ReportType.VISIT_SUMMARY,
            params={},
        ):
            events_out.append(ev)

        # Terminal event must be 'error'.
        assert any(ev.kind == "error" for ev in events_out)
        run = session._captured["run"]
        assert run.status == "failed"
        assert "boom" in (run.error or "")
        assert run.completed_at is not None

    @pytest.mark.asyncio
    async def test_streams_deltas_in_order(self, monkeypatch) -> None:
        """Three text deltas → three SSE delta frames, in the same order."""
        session = _fake_session_with_run()
        events = [
            _msg_start(),
            _text_delta("one "),
            _text_delta("two "),
            _text_delta("three"),
            _msg_delta(),
            _msg_stop(),
        ]
        _patch_anthropic_client(monkeypatch, events)

        service = ReportService()
        deltas: list[str] = []
        async for ev in service.generate(
            session,
            agency_id=uuid.uuid4(),
            user_id=uuid.uuid4(),
            report_type=ReportType.VISIT_SUMMARY,
            params={},
        ):
            if ev.kind == "delta":
                deltas.append(ev.delta or "")

        assert deltas == ["one ", "two ", "three"]
        run = session._captured["run"]
        assert run.narrative == "one two three"


# --------------------------------------------------------------------------
# Exports
# --------------------------------------------------------------------------
class TestRenderCsv:
    def test_basic_table(self) -> None:
        aggregate = {
            "per_caregiver": [
                {"staff_id": "a", "visits_in_window": 3, "hours_in_window": 4.5},
                {"staff_id": "b", "visits_in_window": 1, "hours_in_window": 2.0},
            ],
        }
        out = render_csv(aggregate)
        assert out.startswith(b"staff_id,visits_in_window,hours_in_window")
        rows = list(csv.reader(io.StringIO(out.decode("utf-8"))))
        assert rows[0] == ["staff_id", "visits_in_window", "hours_in_window"]
        assert rows[1] == ["a", "3", "4.5"]

    def test_falls_back_to_kv_table(self) -> None:
        aggregate = {
            "_data_availability": "limited",
            "as_of": "2026-07-01",
            "data_gaps": ["claims table not built"],  # list — skipped
        }
        rows = list(csv.reader(io.StringIO(render_csv(aggregate).decode("utf-8"))))
        assert rows[0] == ["key", "value"]
        # The list value should NOT appear as a row.
        keys = [r[0] for r in rows[1:]]
        assert "data_gaps" not in keys


class TestRenderXlsx:
    def test_returns_valid_xlsx(self) -> None:
        aggregate = {
            "per_caregiver": [{"staff_id": "x", "visits_in_window": 5}],
        }
        out = render_xlsx(aggregate)
        assert out[:2] == b"PK"  # xlsx is a zip — magic header is "PK"
        # Round-trip via openpyxl to confirm it's actually parseable.
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(out))
        sheet = wb["Report"]
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0] == ("staff_id", "visits_in_window")
        assert rows[1] == ("x", 5)


class TestRenderPdf:
    def test_returns_pdf_bytes(self) -> None:
        aggregate = {
            "totals": {"visits": 12, "hours_billed": 9.5},
            "_data_availability": "full",
        }
        out = render_pdf(
            narrative="- Three visits yesterday\n- Two were late",
            aggregate=aggregate,
            report_type="visit_summary",
            agency_id=uuid.uuid4(),
        )
        assert out.startswith(b"%PDF-")


class TestGetArtifact:
    def _run(self, **overrides: Any) -> Any:
        """Build a fake ReportRun for `get_artifact`."""
        run = MagicMock()
        run.aggregate_payload = {
            "per_caregiver": [{"staff_id": "a", "visits_in_window": 3}],
        }
        run.narrative = "Sample narrative"
        run.report_type = "visit_summary"
        run.agency_id = uuid.uuid4()
        run.id = uuid.uuid4()
        for k, v in overrides.items():
            setattr(run, k, v)
        return run

    def test_csv(self) -> None:
        out = get_artifact(self._run(), fmt="csv")
        assert out.startswith(b"staff_id")

    def test_xlsx(self) -> None:
        out = get_artifact(self._run(), fmt="xlsx")
        assert out[:2] == b"PK"

    def test_pdf(self) -> None:
        out = get_artifact(self._run(), fmt="pdf")
        assert out.startswith(b"%PDF-")

    def test_unsupported_format_raises(self) -> None:
        with pytest.raises(ValueError):
            get_artifact(self._run(), fmt="docx")


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
class TestFlattenAggregate:
    def test_picks_first_list_of_dicts(self) -> None:
        agg = {
            "_data_availability": "full",  # scalar — not used
            "by_status": {"a": 1},  # dict — not used
            "per_caregiver": [{"x": 1}, {"x": 2}],  # list-of-dicts — picked
        }
        headers, rows = _flatten_aggregate_for_table(agg)
        assert headers == ["x"]
        assert rows == [[1], [2]]

    def test_kv_fallback_when_no_list(self) -> None:
        agg = {"a": 1, "b": "two", "_data_availability": "limited"}
        headers, rows = _flatten_aggregate_for_table(agg)
        assert headers == ["key", "value"]
        keys = {row[0] for row in rows}
        assert {"a", "b", "_data_availability"} <= keys


class TestComputeCostUsd:
    def test_sonnet_4_5(self) -> None:
        # 1M input + 1M output = $3 + $15 = $18
        cost = _compute_cost_usd("claude-sonnet-4-5", 1_000_000, 1_000_000)
        assert cost == pytest.approx(18.0)

    def test_haiku_4_5(self) -> None:
        # 1M input + 1M output = $0.80 + $4 = $4.80
        cost = _compute_cost_usd("claude-haiku-4-5", 1_000_000, 1_000_000)
        assert cost == pytest.approx(4.80)

    def test_unknown_model_uses_default(self) -> None:
        # Unknown model falls back to Sonnet pricing.
        cost = _compute_cost_usd("claude-unknown", 1_000_000, 0)
        assert cost == pytest.approx(3.0)


__all__ = []  # pytest discovers tests by name
